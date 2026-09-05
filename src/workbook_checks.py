from pathlib import Path
import re

from openpyxl import load_workbook


GALLONS_PER_CUBIC_FOOT = 7.48052


def is_run_id(value):
    """Return True only for production Run IDs such as R-101."""
    return bool(re.fullmatch(r"R-\d+", str(value).strip())) if value is not None else False


def load_workbook_pair(path):
    """Load one copy with formulas and one copy with cached Excel values."""
    path = Path(path)
    formula_book = load_workbook(path, data_only=False)
    value_book = load_workbook(path, data_only=True)
    return formula_book, value_book


def parse_depth_to_inches(value):
    """Convert a Field Notes depth such as '2.5 ft' or '30 in' to inches."""
    if value is None:
        raise ValueError("Depth value is blank.")

    text = str(value).strip().lower()
    match = re.fullmatch(r"(-?\d+(?:\.\d+)?)\s*(in|inch|inches|ft|foot|feet)", text)

    if not match:
        raise ValueError(f"Could not understand depth value: {value!r}")

    number = float(match.group(1))
    unit = match.group(2)

    if unit in {"ft", "foot", "feet"}:
        return number * 12

    return number


def field_note_depth_units_are_inches(path):
    """
    Check Defect 1.

    Every populated depth in Field Notes should use inches.
    Returns (passed, problems).
    """
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Field Notes"]

    problems = []

    for row in range(4, sheet.max_row + 1):
        run_id = sheet.cell(row, 1).value
        depth = sheet.cell(row, 5).value

        if not is_run_id(run_id):
            continue

        text = str(depth).strip().lower()

        if not re.fullmatch(r"-?\d+(?:\.\d+)?\s*(in|inch|inches)", text):
            problems.append(
                f"{run_id}: Field Notes depth is {depth!r}; expected inches."
            )

    return len(problems) == 0, problems


def tank_fill_matches_field_notes(path, tolerance=1e-9):
    """
    Check Defect 2.

    Field Notes measurements are converted to inches and compared with
    the Fill Depth (in) value on the Tank Fill worksheet.
    Returns (passed, problems).
    """
    workbook = load_workbook(path, data_only=False)

    field_notes = workbook["Field Notes"]
    tank_fill = workbook["Tank Fill"]

    source_depths = {}

    for row in range(4, field_notes.max_row + 1):
        run_id = field_notes.cell(row, 1).value
        depth = field_notes.cell(row, 5).value

        if is_run_id(run_id):
            source_depths[run_id] = parse_depth_to_inches(depth)

    problems = []

    for row in range(6, tank_fill.max_row + 1):
        run_id = tank_fill.cell(row, 1).value

        if not is_run_id(run_id):
            continue

        if run_id not in source_depths:
            problems.append(f"{run_id}: no matching Field Notes record.")
            continue

        tank_depth = tank_fill.cell(row, 5).value
        expected_depth = source_depths[run_id]

        if tank_depth is None or abs(float(tank_depth) - expected_depth) > tolerance:
            problems.append(
                f"{run_id}: Tank Fill has {tank_depth} in, "
                f"but Field Notes converts to {expected_depth:g} in."
            )

    return len(problems) == 0, problems


def volume_formulas_are_consistent(path, tolerance=0.01):
    """
    Check the calculated-volume rows.

    Each populated run should use:
    Length * Width * (Depth / 12) * 7.48052

    The test checks both the formula text and the cached Excel result.
    Returns (passed, problems).
    """
    formula_book, value_book = load_workbook_pair(path)

    formula_sheet = formula_book["Tank Fill"]
    value_sheet = value_book["Tank Fill"]

    problems = []

    for row in range(6, formula_sheet.max_row + 1):
        run_id = formula_sheet.cell(row, 1).value

        if not is_run_id(run_id):
            continue

        length = formula_sheet.cell(row, 3).value
        width = formula_sheet.cell(row, 4).value
        depth = formula_sheet.cell(row, 5).value
        formula = formula_sheet.cell(row, 7).value
        cached_value = value_sheet.cell(row, 7).value

        expected_formula = f"=C{row}*D{row}*(E{row}/12)*{GALLONS_PER_CUBIC_FOOT}"

        if str(formula).replace(" ", "") != expected_formula.replace(" ", ""):
            problems.append(
                f"{run_id}: G{row} uses {formula!r}; expected {expected_formula!r}."
            )

        expected_value = (
            float(length)
            * float(width)
            * (float(depth) / 12)
            * GALLONS_PER_CUBIC_FOOT
        )

        if cached_value is None or abs(float(cached_value) - expected_value) > tolerance:
            problems.append(
                f"{run_id}: G{row} is {cached_value}; "
                f"expected about {expected_value:.2f} gallons."
            )

    return len(problems) == 0, problems


def total_formula_includes_all_runs(path):
    """
    Check Defect 3.

    TOTAL RECORDED VOLUME should sum the Calculated Volume cells for every
    populated Run ID.
    Returns (passed, problems).
    """
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Tank Fill"]

    run_rows = [
        row
        for row in range(6, sheet.max_row + 1)
        if is_run_id(sheet.cell(row, 1).value)
    ]

    if not run_rows:
        return False, ["No run rows were found in Tank Fill."]

    first_run_row = min(run_rows)
    last_run_row = max(run_rows)

    total_cell = None

    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 6).value == "TOTAL RECORDED VOLUME":
            total_cell = sheet.cell(row, 7)
            break

    if total_cell is None:
        return False, ["Could not find the TOTAL RECORDED VOLUME cell."]

    expected_formula = f"=SUM(G{first_run_row}:G{last_run_row})"
    actual_formula = total_cell.value

    if str(actual_formula).replace(" ", "").upper() != expected_formula.upper():
        return False, [
            f"{total_cell.coordinate} uses {actual_formula!r}; "
            f"expected {expected_formula!r}."
        ]

    return True, []


def run_all_checks(path):
    """Run all workbook checks and return the results in one dictionary."""
    return {
        "field_note_units": field_note_depth_units_are_inches(path),
        "source_vs_tank_fill": tank_fill_matches_field_notes(path),
        "volume_formulas": volume_formulas_are_consistent(path),
        "total_formula": total_formula_includes_all_runs(path),
    }


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Check the tank-fill workbook.")
    parser.add_argument("workbook", help="Path to the .xlsx workbook to check")
    args = parser.parse_args()

    results = run_all_checks(args.workbook)

    for name, (passed, problems) in results.items():
        print(f"{name}: {'PASS' if passed else 'FAIL'}")
        for problem in problems:
            print(f"  - {problem}")
